import sys, io
if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

"""
M1 — ingestor.py
Location: backend/part_a/ingestor.py

Responsibilities:
  - Accept CSV, JSON, XLSX (and TSV, JSONL as bonus formats)
  - Normalize to a pandas DataFrame
  - Return a structured ingest_result dict consumed by:
      → gemini_classifier.py  (uses: df, column_meta, dataset_name)
      → proxy_detection.py    (uses: df, column_meta)

Output dict shape:
{
    "dataset_name": str,               # stem of the filename
    "source_path": str,                # absolute path as given
    "format": str,                     # "csv" | "json" | "jsonl" | "xlsx" | "tsv"
    "df": pd.DataFrame,                # cleaned, normalized DataFrame
    "row_count": int,
    "column_count": int,
    "column_meta": [                   # one entry per column, in order
        {
            "name": str,               # original column name
            "dtype": str,              # pandas dtype string
            "kind": str,               # "categorical" | "numeric" | "datetime" | "other"
            "null_count": int,
            "unique_count": int,
            "sample_values": list      # up to 5 unique non-null values (for Gemini prompt)
        },
        ...
    ],
    "warnings": [str]                  # non-fatal issues encountered during load
}
"""

import os
import json
import pandas as pd
import numpy as np
from pathlib import Path


# ─────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────

def ingest(file_path: str, max_rows: int = 5000) -> dict:
    """
    Load a dataset file and return a normalized ingest_result dict.

    Args:
        file_path : path to the file (CSV, TSV, JSON, JSONL, or XLSX)
        max_rows  : hard cap on rows loaded — keeps downstream bias stats fast.
                    Set to None to load everything (not recommended for large files).

    Returns:
        ingest_result dict (see module docstring for shape)

    Raises:
        ValueError  : unsupported file extension
        FileNotFoundError : path doesn't exist
    """
    path = Path(file_path).resolve()

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    ext = path.suffix.lower().lstrip(".")
    warnings = []

    # ── dispatch to format-specific loader ──
    loaders = {
        "csv":  _load_csv,
        "tsv":  _load_tsv,
        "json": _load_json,
        "jsonl": _load_jsonl,
        "xlsx": _load_xlsx,
        "xls":  _load_xlsx,   # xlrd handles legacy .xls
    }

    if ext not in loaders:
        raise ValueError(
            f"Unsupported file type '.{ext}'. "
            f"Supported: {', '.join(loaders.keys())}"
        )

    df, load_warnings = loaders[ext](path)
    warnings.extend(load_warnings)

    # ── row cap ──
    if max_rows is not None and len(df) > max_rows:
        warnings.append(
            f"Dataset has {len(df)} rows — sampled to {max_rows} for performance. "
            "Set max_rows=None to load all rows."
        )
        df = df.sample(n=max_rows, random_state=42).reset_index(drop=True)

    # ── normalize column names ──
    df, rename_warnings = _normalize_columns(df)
    warnings.extend(rename_warnings)

    # ── build column metadata ──
    column_meta = _build_column_meta(df)

    return {
        "dataset_name": path.stem,
        "source_path": str(path),
        "format": ext,
        "df": df,
        "row_count": len(df),
        "column_count": len(df.columns),
        "column_meta": column_meta,
        "warnings": warnings,
    }


# ─────────────────────────────────────────────────────────────
# Format-specific loaders — each returns (df, warnings)
# ─────────────────────────────────────────────────────────────

def _load_csv(path: Path) -> tuple[pd.DataFrame, list]:
    warnings = []
    # Try UTF-8 first, fall back to latin-1 for legacy census-style files (UCI Adult uses latin-1)
    for encoding in ("utf-8", "latin-1"):
        try:
            df = pd.read_csv(path, encoding=encoding)
            if encoding != "utf-8":
                warnings.append(f"File is not UTF-8 — loaded with '{encoding}' encoding.")
            return df, warnings
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {path.name} with UTF-8 or latin-1 encoding.")


def _load_tsv(path: Path) -> tuple[pd.DataFrame, list]:
    warnings = []
    for encoding in ("utf-8", "latin-1"):
        try:
            df = pd.read_csv(path, sep="\t", encoding=encoding)
            if encoding != "utf-8":
                warnings.append(f"File is not UTF-8 — loaded with '{encoding}' encoding.")
            return df, warnings
        except UnicodeDecodeError:
            continue
    raise ValueError(f"Could not decode {path.name} with UTF-8 or latin-1 encoding.")


def _load_json(path: Path) -> tuple[pd.DataFrame, list]:
    """
    Handles three common JSON shapes:
      1. Array of objects   → [{...}, {...}]          (most common)
      2. records orient     → {"columns":[], "data":[]} (pandas default export)
      3. Single object      → {...}                    (wrap as single-row df)
    """
    warnings = []
    with open(path, "r", encoding="utf-8") as f:
        raw = json.load(f)

    if isinstance(raw, list):
        if len(raw) == 0:
            raise ValueError("JSON file contains an empty array — nothing to ingest.")
        df = pd.DataFrame(raw)

    elif isinstance(raw, dict):
        # pandas records-orient export has "columns" and "data" keys
        if "columns" in raw and "data" in raw:
            df = pd.DataFrame(raw["data"], columns=raw["columns"])
        else:
            # Treat as a single row
            warnings.append("JSON is a single object — loaded as a one-row DataFrame.")
            df = pd.DataFrame([raw])
    else:
        raise ValueError(f"Unexpected JSON root type: {type(raw).__name__}. Expected list or object.")

    return df, warnings


def _load_jsonl(path: Path) -> tuple[pd.DataFrame, list]:
    """One JSON object per line."""
    warnings = []
    records = []
    with open(path, "r", encoding="utf-8") as f:
        for i, line in enumerate(f):
            line = line.strip()
            if not line:
                continue
            try:
                records.append(json.loads(line))
            except json.JSONDecodeError as e:
                warnings.append(f"Skipped malformed JSONL line {i + 1}: {e}")

    if not records:
        raise ValueError("JSONL file produced no valid records.")

    df = pd.DataFrame(records)
    return df, warnings


def _load_xlsx(path: Path) -> tuple[pd.DataFrame, list]:
    """Load the first sheet of an XLSX or legacy XLS file."""
    warnings = []
    ext = path.suffix.lower()

    try:
        if ext == ".xls":
            df = pd.read_excel(path, engine="xlrd")
        else:
            df = pd.read_excel(path, engine="openpyxl")
    except Exception as e:
        raise ValueError(f"Could not read Excel file '{path.name}': {e}")

    # If file has multiple sheets, we only take the first — warn about it
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True)
        if len(wb.sheetnames) > 1:
            warnings.append(
                f"Excel file has {len(wb.sheetnames)} sheets "
                f"({', '.join(wb.sheetnames)}) — only the first sheet was loaded."
            )
    except Exception:
        pass  # Non-critical — don't fail the whole ingest for this

    return df, warnings


# ─────────────────────────────────────────────────────────────
# Normalization helpers
# ─────────────────────────────────────────────────────────────

def _normalize_columns(df: pd.DataFrame) -> tuple[pd.DataFrame, list]:
    """
    Strip whitespace from column names.
    Deduplicate columns that become identical after stripping (rare but possible).
    Does NOT lowercase or snake_case — preserves original names for Gemini classification.
    """
    warnings = []
    original_cols = list(df.columns)
    stripped_cols = [str(c).strip() for c in original_cols]

    if stripped_cols != original_cols:
        warnings.append("Stripped leading/trailing whitespace from column names.")

    # Deduplicate: if two columns become the same after strip, suffix with _2, _3, ...
    seen = {}
    deduped = []
    for col in stripped_cols:
        if col in seen:
            seen[col] += 1
            new_name = f"{col}_{seen[col]}"
            warnings.append(f"Duplicate column name '{col}' renamed to '{new_name}'.")
            deduped.append(new_name)
        else:
            seen[col] = 1
            deduped.append(col)

    df.columns = deduped

    # Strip whitespace from string cell values — common in CSVs like UCI Adult
    # include both "object" (pandas 2) and "str" (pandas 3 StringDtype)
    str_cols = df.select_dtypes(include=["object", "str"]).columns
    for col in str_cols:
        df[col] = df[col].str.strip() if hasattr(df[col], "str") else df[col]

    return df, warnings


def _infer_kind(series: pd.Series) -> str:
    """
    Classify a column into one of four semantic kinds:
      - "categorical" : object dtype, or numeric with few unique values (≤ 20)
      - "numeric"     : int or float with many unique values
      - "datetime"    : datetime dtype
      - "other"       : anything else (bool, complex, etc.)
    """
    dtype = series.dtype

    if pd.api.types.is_datetime64_any_dtype(dtype):
        return "datetime"

    if pd.api.types.is_bool_dtype(dtype):
        return "categorical"

    # pandas 2: object dtype for strings; pandas 3: pd.StringDtype()
    if pd.api.types.is_object_dtype(dtype) or pd.api.types.is_string_dtype(dtype):
        return "categorical"

    if pd.api.types.is_numeric_dtype(dtype):
        n_unique = series.nunique(dropna=True)
        # Small-cardinality numerics (e.g. encoded 0/1 or age groups) → categorical
        if n_unique <= 20:
            return "categorical"
        return "numeric"

    return "other"


def _build_column_meta(df: pd.DataFrame) -> list:
    """
    Build the column_meta list — one dict per column.
    sample_values are fed directly into the Gemini classification prompt.
    """
    meta = []
    for col in df.columns:
        series = df[col]
        non_null = series.dropna()

        # Up to 5 unique sample values, cast to native Python types for JSON safety
        raw_samples = non_null.unique()[:5].tolist()
        sample_values = []
        for v in raw_samples:
            if isinstance(v, (np.integer,)):
                sample_values.append(int(v))
            elif isinstance(v, (np.floating,)):
                sample_values.append(float(v))
            else:
                sample_values.append(v)

        meta.append({
            "name": col,
            "dtype": str(series.dtype),
            "kind": _infer_kind(series),
            "null_count": int(series.isna().sum()),
            "unique_count": int(series.nunique(dropna=True)),
            "sample_values": sample_values,
        })

    return meta


# ─────────────────────────────────────────────────────────────
# Utility — pretty-print a summary (useful for quick CLI checks)
# ─────────────────────────────────────────────────────────────

def summarize(result: dict) -> None:
    """Print a human-readable summary of an ingest_result dict."""
    print(f"\n{'=' * 55}")
    print(f"  Dataset : {result['dataset_name']}")
    print(f"  Format  : {result['format'].upper()}")
    print(f"  Shape   : {result['row_count']} rows × {result['column_count']} columns")
    print(f"  Path    : {result['source_path']}")
    if result["warnings"]:
        print("\n  ⚠  Warnings:")
        for w in result["warnings"]:
            print(f"     • {w}")
    print(f"\n  {'Column':<25} {'dtype':<12} {'kind':<12} {'nulls':>6} {'unique':>7}")
    print(f"  {'-' * 65}")
    for col in result["column_meta"]:
        print(
            f"  {col['name']:<25} {col['dtype']:<12} {col['kind']:<12} "
            f"{col['null_count']:>6} {col['unique_count']:>7}"
        )
    print(f"{'=' * 55}\n")


# ─────────────────────────────────────────────────────────────
# Quick self-test — run directly: python ingestor.py
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import tempfile

    print("Running ingestor self-test...\n")

    # ── Build a tiny synthetic dataset matching UCI Adult column structure ──
    synthetic_data = {
        "age": [39, 50, 38, 53, 28],
        "workclass": ["State-gov", "Self-emp-not-inc", "Private", "Private", "Private"],
        "fnlwgt": [77516, 83311, 215646, 234721, 338409],
        "education": ["Bachelors", "Bachelors", "HS-grad", "11th", "Bachelors"],
        "education-num": [13, 13, 9, 7, 13],
        "marital-status": ["Never-married", "Married-civ-spouse", "Divorced", "Married-civ-spouse", "Married-civ-spouse"],
        "occupation": ["Adm-clerical", "Exec-managerial", "Handlers-cleaners", "Handlers-cleaners", "Prof-specialty"],
        "relationship": ["Not-in-family", "Husband", "Not-in-family", "Husband", "Wife"],
        "race": ["White", "White", "White", "Black", "Black"],
        "gender": ["Male", "Male", "Male", "Male", "Female"],
        "capital-gain": [2174, 0, 0, 0, 0],
        "capital-loss": [0, 0, 0, 0, 0],
        "hours-per-week": [40, 13, 40, 40, 40],
        "native-country": ["United-States", "United-States", "United-States", "United-States", "Cuba"],
        "income": [">50K", "<=50K", "<=50K", "<=50K", "<=50K"],
    }

    df_synthetic = pd.DataFrame(synthetic_data)

    with tempfile.TemporaryDirectory() as tmp:
        # ── TEST CSV ──
        csv_path = os.path.join(tmp, "uci_adult.csv")
        df_synthetic.to_csv(csv_path, index=False)
        result_csv = ingest(csv_path)
        assert result_csv["format"] == "csv"
        assert result_csv["row_count"] == 5
        assert result_csv["column_count"] == 15
        summarize(result_csv)
        print("✅ CSV: PASS")

        # ── TEST TSV ──
        tsv_path = os.path.join(tmp, "uci_adult.tsv")
        df_synthetic.to_csv(tsv_path, sep="\t", index=False)
        result_tsv = ingest(tsv_path)
        assert result_tsv["format"] == "tsv"
        assert result_tsv["row_count"] == 5
        print("✅ TSV: PASS")

        # ── TEST JSON (array of objects) ──
        json_path = os.path.join(tmp, "uci_adult.json")
        df_synthetic.to_json(json_path, orient="records", indent=2)
        result_json = ingest(json_path)
        assert result_json["format"] == "json"
        assert result_json["row_count"] == 5
        print("✅ JSON (records array): PASS")

        # ── TEST JSONL ──
        jsonl_path = os.path.join(tmp, "uci_adult.jsonl")
        with open(jsonl_path, "w") as f:
            for record in df_synthetic.to_dict(orient="records"):
                f.write(json.dumps(record) + "\n")
        result_jsonl = ingest(jsonl_path)
        assert result_jsonl["format"] == "jsonl"
        assert result_jsonl["row_count"] == 5
        print("✅ JSONL: PASS")

        # ── TEST XLSX ──
        xlsx_path = os.path.join(tmp, "uci_adult.xlsx")
        df_synthetic.to_excel(xlsx_path, index=False)
        result_xlsx = ingest(xlsx_path)
        assert result_xlsx["format"] == "xlsx"
        assert result_xlsx["row_count"] == 5
        print("✅ XLSX: PASS")

        # ── TEST column_meta shape ──
        meta = result_csv["column_meta"]
        assert len(meta) == 15
        assert all(k in meta[0] for k in ("name", "dtype", "kind", "null_count", "unique_count", "sample_values"))
        gender_meta = next(m for m in meta if m["name"] == "gender")
        assert gender_meta["kind"] == "categorical"
        age_meta = next(m for m in meta if m["name"] == "age")
        # age has 5 unique values in 5 rows — ≤ 20 → categorical
        assert age_meta["kind"] == "categorical"
        print("✅ column_meta shape and kinds: PASS")

        # ── TEST unsupported format (file must exist to reach the extension check) ──
        parquet_path = os.path.join(tmp, "bad.parquet")
        with open(parquet_path, "w") as f:
            f.write("dummy")
        try:
            ingest(parquet_path)
            assert False, "Should have raised ValueError"
        except ValueError as e:
            assert "Unsupported file type" in str(e)
        print("✅ Unsupported format error: PASS")

        # ── TEST file not found ──
        try:
            ingest("/tmp/does_not_exist.csv")
            assert False, "Should have raised FileNotFoundError"
        except FileNotFoundError:
            pass
        print("✅ FileNotFoundError: PASS")

    print("\n✅ All ingestor tests passed — ready for gemini_classifier.py 🚀")
